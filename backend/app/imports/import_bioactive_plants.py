from pathlib import Path

from openpyxl import load_workbook

from app.database import SessionLocal
from app.models.bioactive import Bioactive
from app.models.bioactive_plant import BioactivePlant
BASE_DIR = Path(__file__).resolve().parents[3]

file_path = BASE_DIR / "datasets" / "Bioactives.xlsx"

wb = load_workbook(file_path)

ws = wb["Bioactives"]

db = SessionLocal()
headers = {}

for cell in ws[1]:
    if cell.value is not None:
        headers[str(cell.value).strip()] = cell.column
bioactive_col = headers["Bioactive"]
plant_col = headers["Chemical Constituents Containing Plants"]
db.query(BioactivePlant).delete()
db.commit()
print("Importing Bioactive Plants...")

current_bioactive = None
count = 0

# Start from row 2 (skip header)
for row in range(2, ws.max_row + 1):

    bioactive_cell = ws.cell(row=row, column=bioactive_col)
    plant_cell = ws.cell(row=row, column=plant_col)

    # If a new bioactive appears, fetch it from DB
    if bioactive_cell.value:
        bioactive_name = str(bioactive_cell.value).strip()

        current_bioactive = (
            db.query(Bioactive)
            .filter(Bioactive.bioactive_name == bioactive_name)
            .first()
        )

    # Skip if we haven't found a valid bioactive yet
    if current_bioactive is None:
        continue

    # Skip empty plant cells
    if not plant_cell.value:
        continue

    plant_name = str(plant_cell.value).strip()

    # Read Excel hyperlink (if present)
    plant_url = None
    if plant_cell.hyperlink:
        plant_url = plant_cell.hyperlink.target

    plant = BioactivePlant(
        bioactive_id=current_bioactive.bioactive_id,
        plant_name=plant_name,
        plant_url=plant_url,
    )

    db.add(plant)
    count += 1

db.commit()
db.close()

print("=" * 50)
print(f"Successfully imported {count} bioactive plants.")
print("=" * 50)